import openpyxl
import os

#保存するファイル名を入力
saved_file_name = "processed_data"#拡張子はいらない

# Excelファイル名とシート名の取得
cwd = os.getcwd()
for file in os.listdir(cwd):
    if file.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        break

#余計な列削除
sheet.delete_cols(3)
sheet.delete_cols(4)

#各列のタイトル指定
sheet['B1'] = 'aria(nm^2)'
sheet['C1'] = 'diameter(nm)'
sheet['D1'] = 'numbeer of particles'
sheet['E1'] = 'particle size(nm)'
sheet['F1'] = 'count'
sheet['G1'] = 'frequency'
sheet['H1'] = 'average'
sheet['I1'] = 'standard deviation'
sheet['J1'] = 'rsd'

#number of particles取得
number_of_rows = sheet.max_row
number_of_particles = number_of_rows - 1 

#ariaの並び替え
#リストにariaの値を格納
arias = []
#リストの並び替え
for row in sheet.iter_rows(min_row=2,max_row=number_of_rows,min_col=2,max_col=2):
    arias.append([cell.value for cell in row])
sorted_arias = sorted(arias)
#並び替えたariaを書き込み
for i in range(1,len(sorted_arias)):
    sheet.cell(i+1,2,value = sorted_arias[i][0])

#diameter書き込み
for i in range(2,number_of_rows):
    aria = sheet.cell(row=i,column=2).value
    sheet.cell(row=i,column=3).value=2*(aria/3.1416)**0.5

#number of particles書き込み
sheet.cell(row=2,column=4).value=number_of_particles

#particle size書き込み
#Ptナノ粒子用なので、ここはサイズに応じてなおすこと
sheet['E2'] = 1
#グラフにする最大粒子径
sheet['E18'] = '>9' #必要に応じて変える 
#2～8.5まで0.5nmずつ追加
for i in range(2,17): #ここも必要に応じて変える
    size = sheet.cell(row=i,column=5).value
    sheet.cell(row=i+1,column=5).value = size + 0.5

#count書き込み
#diameterをリストに格納
size_for_counting = []
for rows in sheet.iter_rows(min_row=2,max_row=number_of_rows,min_col=3,max_col=3):
    for cell in rows:
        size_for_counting.append(cell.value)
#特定のサイズの粒子を抽出し別のリストに格納
for i in range(2,17): #ここも上に合わせて変える
    min_size = sheet.cell(row=i,column=5).value
    max_size = sheet.cell(row=i+1,column=5).value
    specific_size = [i for i in size_for_counting if i >= min_size]
    object_size = [i for i in specific_size if i < max_size]
#リストの要素数をcountに書き込み
    counts = len(object_size)
    sheet.cell(row=i,column=6).value = counts
#8~8.5nmの粒子数の書き込み
hachi = [i for i in size_for_counting if i >= 8] #ここも上に合わせて変える
hachihan = [i for i in hachi if i < 8.5] #ここも上に合わせて変える
hachihankosuu = len(hachihan)
sheet.cell(row=17,column=6).value = hachihankosuu #ここも上に合わせて変える
#9nm以上の粒子数の書き込み
over_sized = [i for i in size_for_counting if i >= 9] #ここも上に合わせて変える
more_than_9 = len(over_sized)
sheet.cell(row=18,column=6).value = more_than_9 #ここも上に合わせて変える

#frequency書き込み
for i in range(2,19):#ここも上に合わせて変える
    counts_for_frequency = sheet.cell(row=i,column=6).value
    sheet.cell(row=i,column=7).value = counts_for_frequency/number_of_particles*100

#average書き込み
sheet.cell(row=2,column=8).value = sum(size_for_counting)/number_of_particles

#standard deviation書き込み
from statistics import stdev
sheet.cell(row=2,column=9).value = stdev(size_for_counting)

#rsd書き込み
average = sheet.cell(row=2,column=8).value
standard_deviation = sheet.cell(row=2,column=9).value
sheet.cell(row=2,column=10).value = standard_deviation/average

#histglamの作製
from openpyxl.chart import Reference
from openpyxl.chart.text import RichText
from openpyxl.chart.label import DataLabelList

#グラフの種類
graph_obj = openpyxl.chart.BarChart()
#グラフのタイトル
graph_obj.title = RichText(title="粒径分布", font= "Segoe UI", sz=1200)
#グラフのフォントサイズ
graph_obj.style  = 12
#高さ
graph_obj.height = 10
#幅
graph_obj.width  = 15

#Y軸設定
graph_obj.y_axis.title = RichText(title="Frequency [%]", font="Segoe UI", sz=1200) 
graph_obj.y_axis.tickLblFont = RichText(font="Segoe UI", sz=1200)
#X軸設定
graph_obj.x_axis.title = RichText(title="Particle size [nm]", font="Segoe UI", sz=1200)
graph_obj.x_axis.tickLblFont = RichText(font="Segoe UI", sz=1200)
#Y軸範囲指定
Y_axis = Reference(sheet, min_col=7, min_row=2, max_col=7, max_row=18) #ここも上に合わせて変える
graph_obj.add_data(Y_axis, titles_from_data=True)
#X軸範囲指定
X_axis = Reference(sheet, min_col=5, min_row=2, max_col=5, max_row=18) #ここも上に合わせて変える
graph_obj.set_categories(X_axis)

#凡例削除
graph_obj.legend = None

#棒間隔設定
graph_obj.gapWidth = 40

#色指定
from openpyxl.drawing.fill import ColorChoice
ser1 = graph_obj.series[0]
ser1.graphicalProperties.solidFill = ColorChoice(prstClr="cornflowerBlue")
ser1.dLbls = DataLabelList(showVal=True)

#グラフの配置場所指定
sheet.add_chart(graph_obj, "I5")

#ファイルの保存
wb.save(saved_file_name + '.xlsx') #保存するファイル名を入れる